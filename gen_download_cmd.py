#!/volume1/@appstore/py3k/usr/local/bin/python3
import requests
import re
import json
import os
import logging
import argparse
import time
import json
from openpyxl import load_workbook

aid_file   = "./aid_done.list"
cid_file   = "./cid_done.list"
down_file  = "./download.list"
json_file  = "./download.json" 
fail_file  = "./fail.list"
excel_file = "./bilibili.xlsx"

pat_filter    = re.compile(r'/|\s|\x10')
pat_av_number = re.compile(r'^av')

def dump_log(args):
    abs_cur_path = os.path.abspath(os.path.expandvars(os.path.curdir))
    if args.debug:
        level = logging.DEBUG
    else :
        level = logging.INFO

    format_sh = '[%(levelname)s] %(message)s'
    format_fh = '%(asctime)s | %(funcName)s() | L%(lineno)s | [%(levelname)s] \n\t%(message)s'
    '''
        %(levelno)s     打印日志级别的数值
        %(levelname)s   打印日志级别名称
        %(pathname)s    打印当前执行程序的路径
        %(filename)s    打印当前执行程序名称
        %(funcName)s    打印日志的当前函数
        %(lineno)d      打印日志的当前行号
        %(asctime)s     打印日志的时间
        %(thread)d      打印线程id
        %(threadName)s  打印线程名称
        %(process)d     打印进程ID
        %(message)s     打印日志信息
    '''

    if args.log_file:

        logging.basicConfig(
            filename = args.log_file, 
            filemode = 'w', 
            datefmt  = "%y-%m-%d %H:%M:%S",
            format   = format_fh, 
            level    = level)

        console = logging.StreamHandler()
        console.setLevel(level)
        console.setFormatter(logging.Formatter(format_sh))
        logging.getLogger('').addHandler(console);
    else:
        logging.basicConfig(
            format = format_sh, 
            level  = level)

def get_args_top():
    parser = argparse.ArgumentParser()
    st = "debug option"
    parser.add_argument('-d',   '--debug', default = False, action='store_true', help=st)

    st = "only print info"
    parser.add_argument('-oi',  '--only_info', default = False, action='store_true', help=st)

    st = "download enable"
    parser.add_argument('-e',   '--enable_download', default = False, action='store_true', help=st)

    st = "download skip"
    parser.add_argument('-s',   '--skip', default = False, action='store_true', help=st)

    st = "gen_download_list"
    parser.add_argument('-g',   '--gen_download_list', default = False, action='store_true', help=st)

    st = "add mid"
    parser.add_argument('-mid', '--add_mid', default = "", help=st)

    st = "add aid"
    parser.add_argument('-aid', '--add_aid', default = "", help=st)

    st = "set output log file"
    parser.add_argument('-l',   '--log_file', default = "", help=st)

    st = "set output dir name"
    parser.add_argument('-o',   '--base_dir', default = "../../../download/video", help=st)

    st = "set delay time"
    parser.add_argument('-dly', '--delay', default = 0, help=st)

    st = "set retry times"
    parser.add_argument('-t',   '--times', default = 0, help=st)

    args = parser.parse_args()
    return args

def youget_download(cmd, log):
    try :
        info = os.system(cmd)
        log.info(info)
    except Exception as e:
        log.info(e)

def gen_download_list(args, log):
    def gen_sheet(ws, aid_list, source):
        skip_cnt = 0
        fail_cnt = 0
        curr_row = 1

        for col in range(1, len(aidHeaderList)):
            ws.cell(row = curr_row, column = col, value = aidHeaderList[col])

        for aid in aid_list :
            log.debug("aid= %s "%(aid))
            curr_row += 1
            #status = ws.cell(row = curr_row, column = 4).value
            
            status = 'done' if str(aid) in aid_done_set else 'todo'
            ws.cell(row = curr_row, column = aidHeaderHash['status'], value = status)
            if status == 'done' and args.skip == True:
                skip_cnt += 1
                continue
            api = "http://api.bilibili.com/view?type=json&appkey=8e9fc618fbd41e28&id=" + str(aid) + "&batch=1"
            log.debug("get video info api is %s "%(api))
            r = requests.get(api)
            j = json.loads(r.text)
            times = 1
            if 'code' in j.keys() :
                log.warning("Download av%s Fail!"%(aid))
                while 'code' in j.keys() and int(args.delay) > 0 and int(args.times) > 0:
                    time.sleep(args.delay)
                    r = requests.get(api)
                    j = json.loads(r.text)
                    log.warning("Download av%s Fail %d time(s)!"%(aid, times))
                    times += 1
                    if times > args.times:
                        break

                fa = open(fail_file, 'a')
                fa.write("Download av%s Fail!\n"%(aid))
                fa.write("\tapi fail   : %s\n"%(api))
                fa.write("\tapi result : %s\n"%(j))
                fa.close()

                fail_cnt += 1
                continue

            up_name = j['author']
            title   = pat_filter.sub('', j['title'])
            pages   = len(j['list'])
            log.info("title = %s"%title)
            #ws.cell(row = curr_row, column = 3 , value = title)
            ws.cell(row = curr_row, column = aidHeaderHash['aid']           , value = aid)
            ws.cell(row = curr_row, column = aidHeaderHash['title']         , value = title)
            ws.cell(row = curr_row, column = aidHeaderHash['author']        , value = up_name)
            ws.cell(row = curr_row, column = aidHeaderHash['page']          , value = pages)
            ws.cell(row = curr_row, column = aidHeaderHash['created_at']    , value = j['created_at'])
            ws.cell(row = curr_row, column = aidHeaderHash['description']   , value = j['description'])
            if pages > 1 :
                ws.row_dimensions.group(curr_row + 1, curr_row + pages, hidden = True)
                
                for e in j['list']:
                    curr_row += 1
                    cid = e['cid']
                    status = 'done' if str(cid) in cid_done_set else 'todo'

                    url = "https://www.bilibili.com/video/av%s/index_%s.html"%(aid, e['page'])
                    if source == 'aid' :
                        output_dir = "%s/%s/%s"%(args.base_dir, 'other', title )
                    else : # source == 'mid'
                        output_dir = "%s/%s/%s"%(args.base_dir, up_name, title )
                    if not os.path.exists(output_dir) :
                        os.makedirs(output_dir) 
                    output_name = pat_filter.sub('', e['part'].strip())
                    ws.cell(row = curr_row, column = aidHeaderHash['aid']           , value = "")
                    ws.cell(row = curr_row, column = aidHeaderHash['cid']           , value = cid)
                    ws.cell(row = curr_row, column = aidHeaderHash['title']         , value = e['part'])
                    ws.cell(row = curr_row, column = aidHeaderHash['author']        , value = up_name)
                    ws.cell(row = curr_row, column = aidHeaderHash['flag']          , value = 'ON')
                    ws.cell(row = curr_row, column = aidHeaderHash['status']        , value = status)
                    ws.cell(row = curr_row, column = aidHeaderHash['page']          , value = e['page'])
                    ws.cell(row = curr_row, column = aidHeaderHash['created_at']    , value = "")
                    ws.cell(row = curr_row, column = aidHeaderHash['description']   , value = "")
                    ws.cell(row = curr_row, column = aidHeaderHash['download_url']  , value = url)
                    ws.cell(row = curr_row, column = aidHeaderHash['output_dir']    , value = output_dir)
                    ws.cell(row = curr_row, column = aidHeaderHash['output_name']   , value = output_name)
                    flag = ws.cell(row = curr_row, column = aidHeaderHash['flag']).value
                    flag = True if flag is not None and str(flag) != 'OFF' else False
                    if status == 'done' and args.skip == True:
                        skip_cnt += 1
                        continue
                    else : #if status != 'done':
                        log.info("append video %10s/%10s -> '%s' by %s"%(aid, cid, title, up_name))
                        download_list.append((url, output_dir, output_name, aid, cid, flag))
            else :
                url = "https://www.bilibili.com/video/av%s"%(aid)
                if source == 'aid' :
                    output_dir  = "%s/%s"%(args.base_dir, 'other')
                else :
                    output_dir  = "%s/%s"%(args.base_dir, up_name)
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir) 
                output_name = pat_filter.sub('', title)
                cid = j['list'][0]['cid']
                ws.cell(row = curr_row, column = aidHeaderHash['cid']           , value = cid)
                ws.cell(row = curr_row, column = aidHeaderHash['flag']          , value = 'ON')
                ws.cell(row = curr_row, column = aidHeaderHash['download_url']  , value = url)
                ws.cell(row = curr_row, column = aidHeaderHash['output_dir']    , value = output_dir)
                ws.cell(row = curr_row, column = aidHeaderHash['output_name']   , value = output_name)

                flag = ws.cell(row = curr_row, column = aidHeaderHash['flag']).value
                flag = True if flag is not None and str(flag) != 'OFF' else False
                if status != 'done':
                    log.info("append video %10s/%10s -> '%s' by %s"%(aid, cid, title, up_name))
                    download_list.append((url, output_dir, output_name, aid, cid, flag))

        return (skip_cnt, fail_cnt, download_list) 
        
    def get_done_set(filename):
        if not os.path.exists(filename) :
            os.system("touch %s"%filename)

        done_set = set()
        fr = open(filename, 'r')
        for line in fr:
            done_set.add(pat_av_number.sub('', line.strip()))
        fr.close()
        return done_set

    def get_mid_list(sheetname):
        ws = wb[sheetname]
        mid_list = []
        for curr_row in range(2, ws.max_row + 1):
            mid    = ws.cell(row = curr_row, column = midHeaderHash['mid']).value
            enable = ws.cell(row = curr_row, column = midHeaderHash['enable']).value
            if mid is not None and str(enable) == 'ON':
                log.debug("mid = %s"%mid)
                mid_list.append(mid)

        log.info("The total number of up(s) is %s"%len(mid_list))
        return ws, mid_list

    def get_aid_list(sheetname):
        ws = wb[sheetname]
        aid_list = []
        for curr_row in range(2, ws.max_row + 1):
            aid = ws.cell(row = curr_row, column = aidHeaderHash['aid']).value
            #log.info("aid = '%s'"%aid)
            if aid is not None :
                aid = pat_av_number.sub('', aid.strip())
                log.debug("aid = '%s'"%aid)
                aid_list.append(aid)
        log.info("The total number of video(s) is %s"%len(aid_list))
        return ws, aid_list

    def initHeaderList(List):
        Hash = {}
        for i in range(1, len(List)):
            e = List[i]
            Hash[e] = i
        return Hash

    download_list = []
    aidHeaderList = ['', 'aid', 'cid', 'title', 'flag', 'status', 'author', 'created_at', 'description', 'page', 'download_url', 'output_dir', 'output_name']
    midHeaderList = ['', 'mid', 'enable', 'author', '投稿数', '已下载', '下载失败', '待下载']
    aidHeaderHash = initHeaderList(aidHeaderList)
    midHeaderHash = initHeaderList(midHeaderList)

    wb = load_workbook(excel_file)

    aid_done_set = get_done_set(aid_file)
    cid_done_set = get_done_set(cid_file)

    ws, aid_list = get_aid_list('aid')
    log.info("aid_list:%s"%aid_list)
    skip_cnt, fail_cnt, download_list = gen_sheet(ws, aid_list, 'aid')
    log.info("skip/fail/total %3d/ %3d/ %3d"%(skip_cnt, fail_cnt, len(aid_list)))

    ws, mid_list = get_mid_list('mid')
    log.info("mid_list:%s"%mid_list)

    curr_row = 1
    for mid in mid_list :
        curr_row += 1
        #midHeaderList = ['', 'mid', 'enable', 'author', '投稿数', '已下载', '下载失败', '待下载']
        #log.info("curr_row = %s"%curr_row)
        todo_cnt = ws.cell(row = curr_row, column = midHeaderHash['待下载']).value
        #todo_cnt = ws.cell(row = curr_row, column = 6).value
        log.info("mid %s ; todo_cnt %s"%(mid, todo_cnt))
        skip = True if todo_cnt == 0 else False
        #skip = True if todo_cnt is not None and todo_cnt != 0 else False
        if skip == True and args.skip:
        #if skip == True :
            log.info("skip %s"%mid)
            continue
        av_dict, av_list, up_name = get_up_vlist(log, mid)
        log.info("up_name %s"%up_name)
        if up_name not in wb.sheetnames:
            ws = wb.create_sheet()
            log.info("create sheet : %s"%up_name)
            ws.title = up_name
        else :
            #wb.remove_sheet(up_name)
            #ws = wb.create_sheet()
            #ws.title = up_name
            ws = wb[up_name]
        skip_cnt, fail_cnt, download_list = gen_sheet(ws, av_list, 'mid')
        av_cnt = len(av_list)

        ws = wb['mid']
        
        ws.cell(row = curr_row, column = midHeaderHash['mid']     , value = mid)
        ws.cell(row = curr_row, column = midHeaderHash['enable']  , value = 'ON')
        ws.cell(row = curr_row, column = midHeaderHash['author']  , value = up_name)
        ws.cell(row = curr_row, column = midHeaderHash['投稿数']  , value = av_cnt)
        ws.cell(row = curr_row, column = midHeaderHash['已下载']  , value = skip_cnt)
        ws.cell(row = curr_row, column = midHeaderHash['下载失败'], value = fail_cnt)
        ws.cell(row = curr_row, column = midHeaderHash['待下载']  , value = av_cnt - skip_cnt - fail_cnt)

        log.info("skip/fail/total %3d/ %3d/ %3d"%(skip_cnt, fail_cnt, av_cnt))
        
    wb.save(excel_file)

    log.info("There are %d video(s) to be downloaded"%len(download_list))

    download_json = json.dumps(download_list)

    fw = open(json_file, 'w')
    fw.write(download_json)
    fw.close()
    log.info("Generate json file %s"%(json_file))

    return download_list

def bilibili_downloader(args, log, download_list = []):
    def refresh_download_list(i):
        fw = open(down_file, 'w')
        total_num = len(download_list)
        for j in range(len(download_list)) :
            url, output_dir, output_name ,aid ,cid, flag = download_list[j]
            if flag == False :
                fw.write("(%d/%d) %-15s %s \n"%(j+1, total_num, "Skip", output_name))
            elif j < i :
                fw.write("(%d/%d) %-15s %s \n"%(j+1, total_num, "Completed", output_name))
            elif j == i :
                fw.write("============== \n")
                fw.write("(%d/%d) %-15s %s \n"%(j+1, total_num, "Download ->", output_name))
                fw.write("============== \n")
            else:
                fw.write("(%d/%d) %-15s %s \n"%(j+1, total_num, "Waiting...", output_name))

        fw.close()

    fr = open(json_file, 'r')
    for line in fr:
        download_json = line
    fr.close()

    download_list = json.loads(download_json)

    refresh_download_list(0)
    for i in range(len(download_list)):
        url, output_dir, output_name, aid ,cid, flag= download_list[i]
        if flag == False:
            log.info("Skip %s/%s"%(output_dir, output_name))
            continue


        output_dir = re.sub(r'\.+$', '', output_dir)
        down_cmd = "you-get --output-dir '%s' --output-filename '%s' %s"%(output_dir, output_name, url)
        log.info(down_cmd)
        if args.only_info :
            info_cmd = "you-get --info %s"%(url)
            log.info(info_cmd)
            log.info(down_cmd)
            continue

        if args.enable_download :
            refresh_download_list(i)
            youget_download(down_cmd, log)
            flag = os.path.isfile(os.path.join(output_dir, "%s.%s"%(output_name, 'flv')))
            if flag == True:
                fa = open(aid_file, 'a')
                fa.write('av%s \n'%(aid))
                fa.close()

                fa = open(cid_file, 'a')
                fa.write('%s \n'%(cid))
                fa.close()
            else :
                fa = open(fail_file, 'a')
                fa.write("Download Fail! av %s/%s %s.flv not found at %s\n"%(aid, cid, output_name, output_dir))
                fa.close()

def get_up_vlist(log, mid):
    cookies = {
    }

    headers = {

    params = (
    )
    av_dict     = {}
    av_list     = []
    page_size   = 100
    page        = 1
    av_cnt      = None
    while True:
        url = "http://space.bilibili.com/ajax/member/getSubmitVideos?mid=%s&pagesize=%s&page=%s"%(mid, page_size, page)
        log.debug("get up info api is %s"%url)
        r   = requests.get(url, headers=headers, params=params, cookies=cookies)
        j   = json.loads(r.text)
        for e in j['data']['vlist'] :
            log.debug("%10s : %s"%(e['aid'], e['title']))
            av_dict[e['aid']] = e['title']
            av_list.append(e['aid'])

        up_name = e['author']

        if av_cnt == None:
            av_cnt  = j['data']['count']
            page    += 1
        elif av_cnt > 0:
            av_cnt  -= page_size
            page    += 1
        else:
            break
    return av_dict, av_list, up_name

def main():
    args = get_args_top()
    dump_log(args)
    log = logging
    if os.path.exists(fail_file):
        os.remove(fail_file)

    if args.gen_download_list :
        gen_download_list(args, log)

    if args.enable_download or args.only_info :
        bilibili_downloader(args, log)

    if args.add_aid is not None:
        wb = load_workbook(excel_file)
        ws = wb['aid']
        ws.cell(row = ws.max_row + 1, column = 1, value = args.add_aid)
        wb.save(excel_file)

    if args.add_mid is not None:
        wb = load_workbook(excel_file)
        ws = wb['mid']
        ws.cell(row = ws.max_row + 1, column = 1, value = args.add_mid)
        wb.save(excel_file)

if __name__ == '__main__':

    main()
